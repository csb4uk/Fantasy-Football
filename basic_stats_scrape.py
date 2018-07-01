

def main():
    from bs4 import BeautifulSoup
    from bs4 import Tag
    import os
    import openpyxl
    import requests

    wb_str = 'C:\Python36\projects\Fantasy Football Projections\Basic Stats\Basic Passing Stats.xlsx'
    if os.path.exists(wb_str):
        wb = openpyxl.load_workbook(wb_str)
    else:
        wb = openpyxl.Workbook()
    msg = "Input year to extract: "
    yr = str(input(msg))
    wks_created = wb.create_sheet('QB ' + yr)
    wp_str = "https://www.pro-football-reference.com/years/" + yr + "/passing.htm"
    html_page = requests.get(wp_str)
    excel_row_counter = 1
    excel_column_counter = 1
    if html_page.status_code == 200:
        print("Page Found")
        soup = BeautifulSoup(html_page.content, 'html.parser')
        table_soup = soup.find_all('table', attrs={'id': 'passing'})
        for table_tag in table_soup:
            if isinstance(table_tag, Tag):
                file = open("html_Table_Basic_Stats.html", 'w')
                file.write(table_tag.prettify())
            thead_soup = table_tag.find('thead')
            th_soup = thead_soup.find_all('th')
            for th_tag in th_soup:
                wks_created.cell(row=excel_row_counter, column=excel_column_counter).value = th_tag.get_text()
                excel_column_counter += 1
            excel_row_counter += 1
            excel_column_counter = 1
            tbody_soup = table_tag.find('tbody')
            tr_soup = tbody_soup.find_all('tr')
            for trs in tr_soup:
                for child in trs.findChildren():
                    try:
                        a_tags = child.find('a')
                        wks_created.cell(row=excel_row_counter, column=excel_column_counter).value = a_tags.get_text()
                    except:
                        wks_created.cell(row=excel_row_counter, column=excel_column_counter).value = child.get_text()
                        excel_column_counter += 1
                excel_row_counter += 1
                excel_column_counter = 1
        wb.save(wb_str)
    else:
        print("Page Not Found")


main()
